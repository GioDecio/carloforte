import pytest
import openpyxl
from carloforte import extract


def _make_scattered_xlsx(path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["B2"] = "Name"
    ws["C2"] = "Dept"
    ws["B3"] = "Alice"
    ws["C3"] = "Eng"
    ws["B4"] = "Bob"
    ws["C4"] = "Mkt"
    ws["A6"] = "Q1"
    ws["B6"] = "Q2"
    ws["C6"] = "Q3"
    ws["A7"] = 10
    ws["B7"] = 20
    ws["C7"] = 30
    wb.save(path)


def test_extract_default_is_csv(tmp_path):
    p = tmp_path / "sales.xlsx"
    _make_scattered_xlsx(p)
    result = extract(str(p))
    assert "# Sales ·" in result


def test_extract_csv_contains_data(tmp_path):
    p = tmp_path / "sales.xlsx"
    _make_scattered_xlsx(p)
    result = extract(str(p), fmt="csv")
    assert "Alice" in result
    assert "Q1" in result


def test_extract_markdown(tmp_path):
    p = tmp_path / "sales.xlsx"
    _make_scattered_xlsx(p)
    result = extract(str(p), fmt="markdown")
    assert "## Sheet: Sales" in result
    assert "| Name | Dept |" in result


def test_extract_json(tmp_path):
    p = tmp_path / "sales.xlsx"
    _make_scattered_xlsx(p)
    result = extract(str(p), fmt="json")
    assert "Sales" in result
    assert "Alice" in result


def test_extract_invalid_format(tmp_path):
    p = tmp_path / "sales.xlsx"
    _make_scattered_xlsx(p)
    with pytest.raises(ValueError, match="Unknown format"):
        extract(str(p), fmt="xml")


def test_extract_specific_sheet(tmp_path):
    p = tmp_path / "sales.xlsx"
    _make_scattered_xlsx(p)
    result = extract(str(p), sheets=["Sales"])
    assert "Sales" in result
