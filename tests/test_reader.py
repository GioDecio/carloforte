import openpyxl
from carloforte._reader import load_workbook_sheets


def _make_simple_xlsx(path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["A2"] = "Alice"
    ws["B2"] = 30
    wb.save(path)


def test_load_single_sheet(tmp_path):
    p = tmp_path / "simple.xlsx"
    _make_simple_xlsx(p)
    result = load_workbook_sheets(str(p))
    assert "Data" in result
    grid = result["Data"]
    assert grid[0] == ["Name", "Age"]
    assert grid[1] == ["Alice", 30]


def test_load_selected_sheet(tmp_path):
    p = tmp_path / "simple.xlsx"
    _make_simple_xlsx(p)
    result = load_workbook_sheets(str(p), sheets=["Data"])
    assert list(result.keys()) == ["Data"]


def test_unknown_sheet_raises(tmp_path):
    p = tmp_path / "simple.xlsx"
    _make_simple_xlsx(p)
    try:
        load_workbook_sheets(str(p), sheets=["Missing"])
        assert False, "Expected KeyError"
    except KeyError:
        pass
