import openpyxl
from ._models import CellValue

CellGrid = list[list[CellValue]]


def load_workbook_sheets(
    path: str,
    sheets: list[str] | None = None,
) -> dict[str, CellGrid]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = sheets if sheets is not None else list(wb.sheetnames)
    result: dict[str, CellGrid] = {}
    for name in names:
        if name not in wb.sheetnames:
            wb.close()
            raise KeyError(f"Sheet '{name}' not found in {path!r}")
        ws = wb[name]
        result[name] = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return result
